"""
migrate_to_cloud.py – Überträgt alle lokalen Daten (Excel + JSON + Fotos) an die Railway-App.

Verwendung (auf dem Mac ausführen, NACH dem Railway-Deployment):
  python3 migrate_to_cloud.py https://DEINE-APP.railway.app DEIN_ADMIN_KEY

Beispiel:
  python3 migrate_to_cloud.py https://tagebuch-production.railway.app mein-geheimer-schluessel

Der ADMIN_KEY muss in Railway als Umgebungsvariable gesetzt sein (gleicher Wert).
"""

import sys
import os
import json
import base64
import urllib.request
import urllib.error
from datetime import datetime

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Tagebuch.xlsx")
FOOD_LOG   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "food_log.json")
FOOD_PHOTOS= os.path.join(os.path.dirname(os.path.abspath(__file__)), "food_photos")
SETTINGS   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")

def lese_tagebuch():
    """Liest alle Einträge aus Tagebuch.xlsx."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ openpyxl nicht installiert: pip3 install openpyxl")
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Einträge"]
    eintraege = []
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        d  = ws.cell(row=row, column=1).value
        st = ws.cell(row=row, column=2).value or 0
        en = ws.cell(row=row, column=3).value or 0
        ko = ws.cell(row=row, column=4).value or 0
        notiz = ws.cell(row=row, column=6).value or ""
        ort   = ws.cell(row=row, column=10).value or ""
        d_str = d.strftime("%d.%m.%Y") if hasattr(d, "strftime") else str(d)
        eintraege.append({
            "datum":    d_str,
            "stimmung": int(st),
            "energie":  int(en),
            "koerper":  int(ko),
            "notiz":    str(notiz),
            "ort":      str(ort),
        })
        row += 1
    print(f"  ✓ {len(eintraege)} Tagebuch-Einträge aus Excel gelesen")
    return eintraege


def lese_food():
    """Liest Food-Log und kodiert Fotos als base64."""
    if not os.path.exists(FOOD_LOG):
        print("  – Kein food_log.json gefunden, überspringe Essen.")
        return []
    with open(FOOD_LOG, encoding="utf-8") as f:
        entries = json.load(f)

    for e in entries:
        foto_name = e.get("foto")
        if foto_name:
            foto_path = os.path.join(FOOD_PHOTOS, foto_name)
            if os.path.exists(foto_path):
                ext = foto_name.rsplit(".", 1)[-1].lower()
                mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                        "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/jpeg")
                with open(foto_path, "rb") as img:
                    b64 = base64.b64encode(img.read()).decode()
                e["foto"] = f"data:{mime};base64,{b64}"
            else:
                e["foto"] = None

    print(f"  ✓ {len(entries)} Essen-Einträge aus JSON gelesen")
    return entries


def lese_settings():
    if not os.path.exists(SETTINGS):
        return {}
    with open(SETTINGS) as f:
        s = json.load(f)
    print(f"  ✓ Einstellungen: {s}")
    return s


def sende_import(url, admin_key, payload):
    """Sendet alle Daten an /admin/import."""
    body = json.dumps(payload).encode("utf-8")
    req  = urllib.request.Request(
        url + "/admin/import",
        data=body,
        headers={
            "Content-Type":  "application/json",
            "X-Admin-Key":   admin_key,
            "Content-Length": str(len(body)),
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read())
            return result
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"❌ HTTP {e.code}: {body}")
        sys.exit(1)
    except Exception as ex:
        print(f"❌ Fehler: {ex}")
        sys.exit(1)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    url       = sys.argv[1].rstrip("/")
    admin_key = sys.argv[2]

    print(f"\n🚀 Starte Migration → {url}\n")

    print("📖 Lese lokale Daten…")
    eintraege = lese_tagebuch()
    food_log  = lese_food()
    settings  = lese_settings()

    payload = {
        "eintraege": eintraege,
        "food_log":  food_log,
        "settings":  settings,
    }

    total_mb = len(json.dumps(payload).encode()) / 1024 / 1024
    print(f"\n📦 Datenmenge: {total_mb:.1f} MB")
    if total_mb > 50:
        print("⚠️  Sehr viele Fotos — das kann ein paar Minuten dauern.")

    print(f"\n📤 Sende Daten an {url}/admin/import …")
    result = sende_import(url, admin_key, payload)

    if result.get("ok"):
        print(f"\n✅ Migration erfolgreich!")
        print(f"   {result.get('eintraege', 0)} Tagebuch-Einträge importiert")
        print(f"   {result.get('food_log', 0)} Essen-Einträge importiert")
        print(f"\n🌐 App aufrufen: {url}")
    else:
        print(f"❌ Fehler: {result}")


if __name__ == "__main__":
    main()
